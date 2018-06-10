import openpyxl
wb = openpyxl.load_workbook('100万身份证信息大全.xlsx')
sheetnames=wb.sheetnames
sheet=wb.worksheets[0]
data=[]
count=1
for row in sheet.iter_rows():
    if row[0].value is not None :
        data.append(row[0].value)
    sheet.cell(row=count,column=1).value=None
    count=count+1
count=1
for row in data:
    sheet.cell(row=count, column=1).value = row
    count=count+1
wb.save("整理后的身份证.xlsx")
