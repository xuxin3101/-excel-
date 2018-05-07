import openpyxl

class data(object) :
    yearweek=0
    cd=0
    storyname=0
    Last_year_sales_tax_withdrawal_amount=0
    Sales_tax_withdrawal_amount=0
    Execution_of_the_estimated_amount=0
    Last_year_budget_amount=0
    zone=''
    area=''
    total_budget=0
    performance=0
    laseyear_performance=0
    rate=0.0
    strrate=''
    laseyearrate=0.0
    strlaseyearrate=''
    def __init__(self):
        self.total_budget=0


#os.chdir('D:/')
wb = openpyxl.load_workbook('资料.xlsx')
sheetnames=wb.sheetnames
print(sheetnames)
sheet=wb.worksheets[3]
first=[]


for row in  sheet.iter_rows():
    flag = False
    for r in first :
        if r.cd==row[1].value :
            flag=True
            if row[7].value == 1 or row[7].value == 24 or row[7].value == 31 :
                r.total_budget=r.total_budget+row[5].value
                r.performance=r.performance+row[3].value
                r.laseyear_performance=r.laseyear_performance+row[4].value
                if r.total_budget != 0 :
                    r.rate=r.performance/r.total_budget
                    r.strrate='%.1f%%' % (r.rate * 100)
                else:
                    r.rate=100
                    r.strrate='-'

                if r.laseyear_performance != 0 :
                    r.laseyearrate=r.performance/r.laseyear_performance
                    r.strlaseyearrate='%.1f%%' % (r.laseyearrate * 100)
                else:
                    r.laseyearrate=100
                    r.strlaseyearrate='-'
    if flag is False :
        item=data()
        item.yearweek=row[0].value
        item.cd=row[1].value
        item.storyname=row[2].value
        item.area=row[10].value
        item.total_budget=row[5].value
        item.performance=row[3].value

        item.laseyear_performance=row[4].value
        if item.total_budget != 0 and item.area!='NaN':
            first.append(item)



first.sort(key=lambda x:(x.laseyearrate,x.rate),reverse=True)
for r in first :
        print (str(r.cd)+'   '+r.storyname+'    '+r.area+'  '+str(r.total_budget)+'    '+str(r.performance)+'  '+str(r.laseyear_performance)+' '+r.strrate+'   '+r.strlaseyearrate)

wb=openpyxl.load_workbook('答案.xlsx')
print ( wb.sheetnames)
sheet=wb.worksheets[0]
second=[]
first.pop()
count=3
for r in first :
    sheet.cell(row=count,column=3).value=r.cd
    sheet.cell(row=count,column=4).value=r.storyname
    sheet.cell(row=count, column=5).value = r.area
    sheet.cell(row=count,column=6).value=float(r.total_budget)/1000
    sheet.cell(row=count, column=7).value=float(r.performance)/1000
    sheet.cell(row=count, column=8).value=float(r.laseyear_performance)/1000
    sheet.cell(row=count, column=9).value=r.strrate
    sheet.cell(row=count, column=10).value=r.strlaseyearrate
    count=count+1
wb.save("测试答案.xlsx")
