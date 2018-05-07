# coding=utf-8
import  openpyxl
import os
class data(object) :
    yearweek = 0
    cd = 0
    storyname = 0
    Last_year_sales_tax_withdrawal_amount = 0
    Sales_tax_withdrawal_amount = 0
    Execution_of_the_estimated_amount = 0
    Last_year_budget_amount = 0
    zone = ''
    area = ''
    total_budget = 0
    lastyear_total_budget=0
    performance = 0
    laseyear_performance = 0
    rate = 0.0
    strrate = ''
    laseyearrate = 0.0
    strlaseyearrate = ''
    def __init__(self):
        self=self

os.chdir('d:/')
wb = openpyxl.load_workbook('资料.xlsx')
sheetnames=wb.sheetnames
print(sheetnames)
sheets=wb.sheetnames
sheet=wb.worksheets[3]
zone=[]
area=[]
for row in sheet.iter_rows() :
    flag=False
    for r in zone :
        if r.zone == row[9].value :
            flag=True
            if row[7].value == 1 or row[7].value == 24 or row[7].value == 31 :
                r.total_budget = r.total_budget + row[5].value
                r.performance = r.performance + row[3].value
                r.lastyear_total_budget=r.lastyear_total_budget+row[6].value
                r.laseyear_performance = r.laseyear_performance + row[4].value
                if r.total_budget != 0:
                    r.rate = r.performance / r.total_budget
                    r.strrate = '%f%%' % (r.rate * 100)
                else:
                    r.rate = 100
                    r.strrate = '-'

                if r.laseyear_performance != 0:
                    r.laseyearrate = r.performance / r.laseyear_performance
                    r.strlaseyearrate = '%f%%' % (r.laseyearrate * 100)
                else:
                    r.laseyearrate = 100
                    r.strlaseyearrate = '-'
    if flag is False :
        if row[9].value == 'トライウェル' or row[9].value=='天草':
            continue

        item = data()
        item.yearweek = row[0].value
        item.zone=row[9].value
        item.cd = row[1].value
        item.storyname = row[2].value
        item.lastyear_total_budget=row[6].value
        item.area = row[10].value
        item.total_budget = row[5].value
        item.performance = row[3].value

        item.laseyear_performance = row[4].value
        if item.total_budget != 0 and item.area != 'NaN':
            zone.append(item)
zone.sort(key=lambda x:(x.laseyearrate,x.rate),reverse=True)

zone.pop()
for r in zone :
        print (r.zone+' '+str(r.total_budget)+'    '+str(r.laseyear_performance)+' '+str(r.performance) +'  '+r.strrate+'   '+r.strlaseyearrate)


wb_answer=openpyxl.load_workbook('答案.xlsx')
sheet_answer=wb_answer.worksheets[1]
count=3

for r in zone :
    sheet_answer.cell(row=count,column=3).value=count-2
    sheet_answer.cell(row=count, column=4).value = r.zone
    sheet_answer.cell(row=count,column=5).value=float(r.total_budget)/1000
    sheet_answer.cell(row=count, column=7).value=float(r.performance)/1000
    sheet_answer.cell(row=count, column=6).value=float(r.laseyear_performance)/1000
    sheet_answer.cell(row=count, column=8).value=r.strrate
    sheet_answer.cell(row=count, column=9).value=r.strlaseyearrate
    count=count+1


sheets=wb.sheetnames
sheet=wb.worksheets[3]

for row in sheet.iter_rows():
    flag = False
    for r in area:
        if r.area == row[10].value:
            flag = True
            if row[7].value == 1 or row[7].value == 24 or row[7].value == 31:
                r.total_budget = r.total_budget + row[5].value
                r.performance = r.performance + row[3].value
                r.lastyear_total_budget = r.lastyear_total_budget + row[6].value
                r.laseyear_performance = r.laseyear_performance + row[4].value
                if r.total_budget != 0:
                    r.rate = r.performance / r.total_budget
                    r.strrate = '%f%%' % (r.rate * 100)
                else:
                    r.rate = 100
                    r.strrate = '-'

                if r.laseyear_performance != 0:
                    r.laseyearrate = r.performance / r.laseyear_performance
                    r.strlaseyearrate = '%f%%' % (r.laseyearrate * 100)
                else:
                    r.laseyearrate = 100
                    r.strlaseyearrate = '-'
    if flag is False:
        if row[9].value == 'トライウェル' or row[9].value=='天草':
            continue
        item = data()
        item.yearweek = row[0].value
        item.area=row[10].value
        item.cd = row[1].value
        item.storyname = row[2].value
        item.lastyear_total_budget = row[6].value
        item.total_budget = row[5].value
        item.performance = row[3].value
        item.laseyear_performance = row[4].value
        if item.total_budget != 0 and item.area != 'NaN':
            area.append(item)
area.sort(key=lambda x: (x.laseyearrate, x.rate), reverse=True)

for r in area :
        print (r.area+' '+str(r.total_budget)+'    '+str(r.laseyear_performance)+' '+str(r.performance) +'  '+r.strrate+'   '+r.strlaseyearrate)


area.pop()
count1=1
for r in area :
    sheet_answer.cell(row=count,column=3).value=count1
    sheet_answer.cell(row=count, column=4).value = r.area
    sheet_answer.cell(row=count,column=5).value=float(r.total_budget)/1000
    sheet_answer.cell(row=count, column=7).value=float(r.performance)/1000
    sheet_answer.cell(row=count, column=6).value=float(r.laseyear_performance)/1000
    sheet_answer.cell(row=count, column=8).value=r.strrate
    sheet_answer.cell(row=count, column=9).value=r.strlaseyearrate
    count=count+1
    count1=count1+1




wb_answer.save("测试答案.xlsx")